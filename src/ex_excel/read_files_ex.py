from openpyxl import load_workbook
import pandas as pd

def read_single_sheet(workbook, sheet_name):

        # Загрузить лист из рабочей книги
        sheet = workbook[sheet_name]

        # Прочитать необработанные данные, включая заголовки
        sheet_data_raw = sheet.values

        # Отделить заголовки в отдельную переменную
        columns = next(sheet_data_raw)[0:]

        # Создать DataFrame на основе второй и последующих строк данных с заголовками в качестве названий столбцов и вернуть его
        return pd.DataFrame(sheet_data_raw, columns=columns)


def read_multiple_sheets(file_path):

    # Загрузить рабочую книгу
    workbook = load_workbook(file_path)

    # Получить список всех названий листов в рабочей книге
    sheet_names = workbook.sheetnames

    # Перебрать названия листов, загрузить данные для каждого и объединить их в один DataFrame
    return pd.concat([read_single_sheet(workbook=workbook, sheet_name=sheet_name) for sheet_name in sheet_names], ignore_index=True)

# Определить путь к файлу и названия листов
file_path = f'F:\Языки\Python\Excel_Python_ R\data\iris_data.xlsx'

# Прочитать данные с нескольких листов
consolidated_data = read_multiple_sheets(file_path)

# Вывести объединенные данные
print(consolidated_data.head())