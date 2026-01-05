import openpyxl
import pandas as pd

def read_excel_sheet(file_path, sheet_name):
    """
    Читает данные с указанного листа Excel-файла и возвращает их в виде pandas DataFrame.
    Предполагается, что первая строка содержит заголовки столбцов.

    :param file_path: Путь к Excel-файлу.
    :param sheet_name: Имя листа для чтения.
    :return: pandas DataFrame с данными из листа.
    """
    # Загрузить рабочую книгу
    wb = openpyxl.load_workbook(file_path)

    # Выбрать лист
    sheet = wb[sheet_name]

    # Извлечь значения (включая заголовки)
    sheet_data_raw = sheet.values

    # Разделить заголовки в отдельную переменную
    header = next(sheet_data_raw)[0:]

    # Создать DataFrame на основе оставшихся строк данных с заголовками в качестве названий столбцов
    df = pd.DataFrame(sheet_data_raw, columns=header)

    # Закрыть рабочую книгу для освобождения ресурсов
    wb.close()

    return df


def read_all_excel_sheets(file_path):
    """
    Читает данные со всех листов Excel-файла и объединяет их в один pandas DataFrame.
    Предполагается, что на всех листах структура данных одинакова (одинаковые заголовки в первой строке).

    :param file_path: Путь к Excel-файлу.
    :return: pandas DataFrame с объединёнными данными из всех листов.
    """
    # Загрузить рабочую книгу
    wb = openpyxl.load_workbook(file_path)

    # Получить список всех имён листов
    sheet_names = wb.sheetnames

    # Список для хранения DataFrame'ов с каждого листа
    dataframes = []

    # Пройти по каждому листу
    for sheet_name in sheet_names:
        # Выбрать лист
        sheet = wb[sheet_name]

        # Извлечь значения (включая заголовки)
        sheet_data_raw = sheet.values

        # Разделить заголовки в отдельную переменную
        header = next(sheet_data_raw)[0:]

        # Создать DataFrame на основе оставшихся строк данных с заголовками в качестве названий столбцов
        df = pd.DataFrame(sheet_data_raw, columns=header)

        # Добавить DataFrame в список
        dataframes.append(df)

    # Закрыть рабочую книгу для освобождения ресурсов
    wb.close()

    # Объединить все DataFrame'ы в один, игнорируя исходные индексы
    consolidated_df = pd.concat(dataframes, ignore_index=True)

    return consolidated_df